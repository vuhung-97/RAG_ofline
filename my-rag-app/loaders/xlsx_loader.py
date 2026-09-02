import openpyxl
from loaders.base import BaseLoader

class XlsxLoader(BaseLoader):
    """SRP: Trích xuất văn bản từ Excel (.xlsx) dưới dạng bảng Markdown."""

    def load(self, file_path: str) -> str:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        text_parts = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text_parts.append(f"### Sheet: {sheet_name}")

            for row in sheet.iter_rows(values_only=True):
                # Bỏ qua dòng rỗng
                if not any(row):
                    continue
                row_cells = [str(cell).strip() if cell is not None else "" for cell in row]
                text_parts.append("| " + " | ".join(row_cells) + " |")

        return "\n".join(text_parts)
